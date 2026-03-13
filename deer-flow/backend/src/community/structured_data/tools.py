import csv
import json
from pathlib import Path
from statistics import mean, median

from langchain.tools import tool

from src.config.paths import VIRTUAL_PATH_PREFIX, get_paths

_STRUCTURED_EXTENSIONS = {".csv", ".xlsx", ".xls"}
_DOCUMENT_EXTENSIONS = {".pdf", ".doc", ".docx"}
_TEXT_EXTENSIONS = {".md", ".txt"}


def _detect_column_type(values: list[str]) -> str:
    """Heuristically detect column type from sample values."""
    numeric_count = 0
    date_count = 0
    total = 0
    for v in values:
        v = v.strip()
        if not v:
            continue
        total += 1
        # Check numeric (strip currency/percent symbols)
        cleaned = v.replace(",", "").replace(" ", "")
        for prefix in ("$", "€", "£", "¥"):
            cleaned = cleaned.lstrip(prefix)
        cleaned = cleaned.rstrip("%")
        try:
            float(cleaned)
            numeric_count += 1
            continue
        except ValueError:
            pass
        # Simple date heuristics
        if any(sep in v for sep in ["-", "/"]) and any(c.isdigit() for c in v):
            parts = v.replace("/", "-").split("-")
            if len(parts) >= 2 and all(p.strip().isdigit() for p in parts[:2]):
                date_count += 1
    if total == 0:
        return "text"
    if numeric_count / total >= 0.7:
        return "numeric"
    if date_count / total >= 0.7:
        return "date"
    return "text"


def _parse_numeric_value(v: str) -> float | None:
    """Try to parse a string as a number."""
    cleaned = v.replace(",", "").replace(" ", "").strip()
    for prefix in ("$", "€", "£", "¥"):
        cleaned = cleaned.lstrip(prefix)
    cleaned = cleaned.rstrip("%")
    try:
        return float(cleaned)
    except (ValueError, TypeError):
        return None


def _read_csv(file_path: Path, max_rows: int) -> dict:
    """Read a CSV file and return structured data."""
    with open(file_path, encoding="utf-8", errors="ignore", newline="") as f:
        # Sniff delimiter
        sample = f.read(4096)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample)
            reader = csv.reader(f, dialect)
        except csv.Error:
            reader = csv.reader(f)

        rows: list[list[str]] = []
        total_rows = 0
        for row in reader:
            total_rows += 1
            if total_rows == 1:
                # Always store header row
                rows.append(row)
                continue
            # Store up to max_rows data rows in memory, but keep counting beyond that
            if len(rows) <= max_rows:  # rows includes header, so len(rows) == max_rows + 1 means we've stored max_rows data rows
                rows.append(row)

    if not rows:
        return {"columns": [], "column_types": [], "row_count": 0, "data": [], "summary_stats": {}}

    columns = rows[0]
    data_rows = rows[1 : max_rows + 1]
    # total_rows counts all rows including header; row_count should be data rows only
    row_count = max(total_rows - 1, 0)

    # Detect column types
    col_values = {i: [] for i in range(len(columns))}
    for row in data_rows:
        for i, val in enumerate(row):
            if i < len(columns):
                col_values[i].append(val)

    column_types = [_detect_column_type(col_values.get(i, [])) for i in range(len(columns))]

    # Compute summary stats for numeric columns
    summary_stats = {}
    for i, (col_name, col_type) in enumerate(zip(columns, column_types)):
        if col_type == "numeric":
            nums = [n for v in col_values.get(i, []) if (n := _parse_numeric_value(v)) is not None]
            if nums:
                summary_stats[col_name] = {
                    "min": min(nums),
                    "max": max(nums),
                    "mean": round(mean(nums), 2),
                    "median": round(median(nums), 2),
                    "count": len(nums),
                }

    return {
        "columns": columns,
        "column_types": column_types,
        "row_count": row_count,
        "data": [row for row in data_rows],
        "summary_stats": summary_stats,
    }


def _read_xlsx(file_path: Path, sheet_name: str, max_rows: int) -> dict:
    """Read an Excel file and return structured data."""
    try:
        import openpyxl
    except ImportError:
        return {"error": "openpyxl is not installed. Install it with: uv add openpyxl"}

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([str(cell) if cell is not None else "" for cell in row])
        if len(rows) > max_rows + 1:
            break

    wb.close()

    if not rows:
        return {"columns": [], "column_types": [], "row_count": 0, "data": [], "summary_stats": {}, "sheet_name": ws.title}

    columns = rows[0]
    data_rows = rows[1 : max_rows + 1]

    # Detect column types
    col_values = {i: [] for i in range(len(columns))}
    for row in data_rows:
        for i, val in enumerate(row):
            if i < len(columns):
                col_values[i].append(val)

    column_types = [_detect_column_type(col_values.get(i, [])) for i in range(len(columns))]

    # Compute summary stats for numeric columns
    summary_stats = {}
    for i, (col_name, col_type) in enumerate(zip(columns, column_types)):
        if col_type == "numeric":
            nums = [n for v in col_values.get(i, []) if (n := _parse_numeric_value(v)) is not None]
            if nums:
                summary_stats[col_name] = {
                    "min": min(nums),
                    "max": max(nums),
                    "mean": round(mean(nums), 2),
                    "median": round(median(nums), 2),
                    "count": len(nums),
                }

    return {
        "columns": columns,
        "column_types": column_types,
        "row_count": len(data_rows),
        "data": data_rows,
        "summary_stats": summary_stats,
        "sheet_name": ws.title,
    }


@tool("extract_structured_data", parse_docstring=True)
def extract_structured_data_tool(thread_id: str, filename: str, sheet_name: str = "", max_rows: int = 500) -> str:
    """Extract structured data (tables, columns, statistics) from uploaded spreadsheets or CSVs.

    Preserves table structure that markdown conversion loses. Returns columns,
    types, data rows, and summary statistics for numeric columns.

    Args:
        thread_id: The current thread id.
        filename: Name of the uploaded file to extract (e.g. "data.csv", "report.xlsx").
        sheet_name: For Excel files, the sheet to read (empty string uses the active sheet).
        max_rows: Maximum number of data rows to return (default 500).
    """
    uploads_dir = get_paths().sandbox_uploads_dir(thread_id)
    file_path = uploads_dir / filename

    if not file_path.exists():
        return json.dumps(
            {
                "ok": False,
                "message": f"File not found: {filename}",
                "uploads_dir": f"{VIRTUAL_PATH_PREFIX}/uploads",
            },
            ensure_ascii=False,
            indent=2,
        )

    suffix = file_path.suffix.lower()
    if suffix not in _STRUCTURED_EXTENSIONS:
        return json.dumps(
            {
                "ok": False,
                "message": f"Unsupported file type: {suffix}. Supported: {', '.join(sorted(_STRUCTURED_EXTENSIONS))}",
            },
            ensure_ascii=False,
            indent=2,
        )

    try:
        if suffix == ".csv":
            result = _read_csv(file_path, max_rows)
        else:
            result = _read_xlsx(file_path, sheet_name, max_rows)
    except Exception as e:
        return json.dumps(
            {"ok": False, "message": f"Error reading {filename}: {e}"},
            ensure_ascii=False,
            indent=2,
        )

    if "error" in result:
        return json.dumps({"ok": False, "message": result["error"]}, ensure_ascii=False, indent=2)

    return json.dumps(
        {
            "ok": True,
            "filename": filename,
            "virtual_path": f"{VIRTUAL_PATH_PREFIX}/uploads/{filename}",
            **result,
        },
        ensure_ascii=False,
        indent=2,
    )


@tool("list_uploaded_data_files", parse_docstring=True)
def list_uploaded_data_files_tool(thread_id: str) -> str:
    """List and categorize uploaded files by type (structured data, documents, text).

    Use this to assess what data is available before choosing extraction tools.

    Args:
        thread_id: The current thread id.
    """
    uploads_dir = get_paths().sandbox_uploads_dir(thread_id)
    uploads_dir.mkdir(parents=True, exist_ok=True)

    categorized: dict[str, list[dict]] = {
        "structured_data": [],
        "document": [],
        "text": [],
        "other": [],
    }

    for file_path in sorted(uploads_dir.iterdir()):
        if not file_path.is_file():
            continue
        suffix = file_path.suffix.lower()
        try:
            size = file_path.stat().st_size
        except OSError:
            size = 0

        entry = {
            "filename": file_path.name,
            "virtual_path": f"{VIRTUAL_PATH_PREFIX}/uploads/{file_path.name}",
            "size_bytes": size,
            "extension": suffix,
        }

        if suffix in _STRUCTURED_EXTENSIONS:
            categorized["structured_data"].append(entry)
        elif suffix in _DOCUMENT_EXTENSIONS:
            categorized["document"].append(entry)
        elif suffix in _TEXT_EXTENSIONS:
            categorized["text"].append(entry)
        else:
            categorized["other"].append(entry)

    total = sum(len(v) for v in categorized.values())
    return json.dumps(
        {
            "ok": True,
            "total_files": total,
            "categories": categorized,
        },
        ensure_ascii=False,
        indent=2,
    )
